# excel_writer.py
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging
from datetime import datetime # <--- IMPORT ADDED
import pathlib # <--- Import pathlib if using Path objects
# Import the specific mappings needed from config using direct imports
from config import KOPF_MAP_DATA_CELLS, POS_COLS_DEFS

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# Define static labels and their approximate positions for Kopf sheet
KOPF_STATIC_LABELS = {
    "A1": "D&M KG", "A2": "Auf den Dorfwiesen 1-5", "A3": "56204 Hillscheid",
    "D2": "Bestellblatt Fehro.AR", "E2": "AV-Import",
    "G1": datetime.now().strftime('%d.%m.%Y'), # Uses datetime
    "A6": "Eingabe Kundennummern", "C6": "Kundenadr.", "E6": "Rechnungsadr.",
    "G6": "Lieferadr.", "I6": "AufBestAdr.",
    "A9": "Bestellinformationen", "C9": "Bestelldatum", "D9": "Auftragsname",
    "E9": "Kunden-Auftrags-Nr.", "F9": "Wunsch-Liefertermin", "G9": "Besteller",
    "A12": "Sonderausführung", "E12": "Hinweistext",
    "A14": "Angaben Verschattungselemente", "C14": "BehangArt", "D14": "Kurbelstange",
    "A16": "Farben Verschattungselemente", "C16": "Behang", "D16": "Anschlagstopfen",
    "E16": "Endleiste", "F16": "Aussenkasten", "G16": "Reviblende", "H16":"Führungsschiene",
    "A18": "Farben Insektenschutzelemente", "C18": "Element", "D18": "Endleiste",
    "E18": "Führungsschiene",
    "A20": "Hinweise zur getroffenen Farbauswahl:",
    "C20": "Farbauswahl Behang nur bei ALU-Rollladen möglich.",
    "C21": "Farbauswahl Reviblende nur bei ALU möglich.",
    "C22": "Farbauswahl Führungsschiene nur bei ALU möglich.",
    "A24": "Bestellblatt_Kopf",
}

def write_to_excel(mapped_data, output_directory, base_filename):
    """
    Writes the mapped data to a new Excel file generated programmatically.
    """
    # Ensure output_directory is a Path object
    if not isinstance(output_directory, pathlib.Path):
        output_directory = pathlib.Path(output_directory)

    try:
        wb = Workbook()
        kopf_sheet = wb.active
        kopf_sheet.title = "Bestellblatt_Kopf"
        pos_sheet = wb.create_sheet("Bestellblatt_Positionen")
        logging.info(f"Created workbook with sheets: {wb.sheetnames}")

        bold_font = Font(bold=True)
        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True) # Added wrap_text
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True) # Added wrap_text
        right_align = Alignment(horizontal="right", vertical="center", wrap_text=True) # Added wrap_text

        logging.info("Writing static labels to Kopf sheet...")
        for cell_ref, label_text in KOPF_STATIC_LABELS.items():
            if label_text:
                try:
                    cell = kopf_sheet[cell_ref]; cell.value = label_text
                    if cell_ref in ["A9", "A12", "A14", "A16", "A18", "A20"]: cell.font = bold_font
                except Exception as e: logging.error(f"Error writing static label '{label_text}' to {cell_ref}: {e}")

        logging.info("Writing dynamic data to Kopf sheet...")
        kopf_values = mapped_data.get("kopf", {})
        if not KOPF_MAP_DATA_CELLS: logging.error("KOPF_MAP_DATA_CELLS empty"); return None
        for key, cell_ref in KOPF_MAP_DATA_CELLS.items():
            value = kopf_values.get(key);
            if value is not None:
                try: cell = kopf_sheet[cell_ref]; cell.value = value; cell.alignment = left_align
                except Exception as e: logging.error(f"Error writing Kopf data key '{key}' to {cell_ref}: {e}")

        logging.info("Writing headers to Positionen sheet...")
        if not POS_COLS_DEFS: logging.error("POS_COLS_DEFS empty"); return None
        try:
            sorted_pos_cols = sorted(POS_COLS_DEFS.items(), key=lambda item: item[1][1])
            for key, (header_text, col_idx) in sorted_pos_cols:
                cell = pos_sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border; cell.fill = grey_fill
        except Exception as e: logging.error(f"Error writing Positionen headers: {e}", exc_info=True); return None

        logging.info("Writing data to Positionen sheet...")
        start_row = 2
        positions = mapped_data.get("positionen", [])
        if not positions: logging.warning("No position data found for Excel sheet.")

        for i, pos_data in enumerate(positions):
            # Skip the last row if "FeBreite_11" and "FeHoehe_12" are empty
            if i == len(positions) - 1 and not pos_data.get("FeBreite_11") and not pos_data.get("FeHoehe_12"):
                logging.info(f"Skipping last row in Excel due to empty 'FeBreite_11' and 'FeHoehe_12'.")
                continue

            row_num = start_row + i
            for key, (header_text, col_idx) in POS_COLS_DEFS.items():
                value = pos_data.get(key)
                if value is not None:
                    cell = pos_sheet.cell(row=row_num, column=col_idx)
                    try:
                        numeric_keys = ["Anzahl_Links_13", "Anzahl_Rechts_14", "FeBreite_11", "FeHoehe_12",
                                        "ISS_Behindertengerecht_40", "ISS_Anzahl_Links_41", "ISS_Anzahl_Rechts_42",
                                        "EinzelteilAnzahl_25"]
                        if key in numeric_keys:
                            try:
                                num_value = float(str(value).replace(',', '.'))
                                cell.value = int(num_value) if num_value.is_integer() else num_value
                                cell.alignment = right_align
                                cell.number_format = '0'
                            except (ValueError, TypeError):
                                cell.value = str(value)
                                cell.alignment = left_align
                        else:
                            cell.value = str(value)
                            cell.alignment = left_align
                        cell.border = thin_border
                    except Exception as cell_e:
                        logging.warning(f"Cell write error R{row_num}C{col_idx}: {cell_e}")

        logging.info("Adjusting column widths...")
        for sheet in [kopf_sheet, pos_sheet]:
             column_widths = {}; max_rows_to_check = 500 # Limit rows for performance
             for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows_to_check)):
                for cell in row:
                    if cell.value:
                        length = len(str(cell.value)); padding = 4 if cell.row == 1 else 2; length += padding
                        column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 0), length)
             for col_letter, width in column_widths.items():
                 adjusted_width = max(width, 8); adjusted_width = min(adjusted_width, 50)
                 sheet.column_dimensions[col_letter].width = adjusted_width

        output_filename = f"{base_filename}.xlsx"
        output_path = output_directory / output_filename # Correct path definition
        logging.info(f"Saving workbook to {output_path}...")
        wb.save(output_path)
        logging.info(f"Successfully generated Excel file: {output_path}")
        return str(output_path)

    except Exception as e:
        logging.error(f"Error generating or writing to Excel: {e}", exc_info=True)
        return None